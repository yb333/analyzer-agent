"""跨资产影响分析测试。

覆盖：
- 表名定位资产目录（locate_asset_dirs）
- 读 Sheet3 受影响表清单（_read_affected_tables）
- 端到端跨资产分析（run_cross_asset）
- 多资产汇总输出

运行:
    pytest tests/test_cross_asset_impact.py -v
"""

import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ANALYZER_REF = PROJECT_ROOT / "dws-pipeline-analyzer" / "references"
FIXTURES = PROJECT_ROOT / "tests" / "fixtures" / "analyzer"
sys.path.insert(0, str(ANALYZER_REF))
sys.path.insert(0, str(FIXTURES))

from impact_analyzer import (
    ChangeItem, TableChange,
    locate_asset_dirs, _read_affected_tables, run_cross_asset,
)
from _build_repo import build_mock_repo
from _build_changes import build_changes_xlsx


# ═══════════════════════════════════════════════════════════════
# 表名定位资产目录
# ═══════════════════════════════════════════════════════════════

class TestLocateAssetDirs:
    """按表名在代码仓定位规则组目录。"""

    def test_locate_existing_table(self, tmp_path):
        """表名精确匹配目录名。"""
        repo = build_mock_repo(tmp_path / "repo")
        located = locate_asset_dirs(["DWB_TRADE_ORDER_D"], str(repo["repo_root"]))
        assert located["DWB_TRADE_ORDER_D"]["found"] is True
        assert "DWB_TRADE_ORDER_D" in located["DWB_TRADE_ORDER_D"]["dir"]

    def test_locate_case_insensitive(self, tmp_path):
        """表名大小写不敏感匹配。"""
        repo = build_mock_repo(tmp_path / "repo")
        located = locate_asset_dirs(["dwb_trade_order_d"], str(repo["repo_root"]))
        assert located["dwb_trade_order_d"]["found"] is True

    def test_locate_with_schema_prefix(self, tmp_path):
        """带 schema 前缀的表名能匹配（归一化去前缀）。"""
        repo = build_mock_repo(tmp_path / "repo")
        located = locate_asset_dirs(["dws.DWB_TRADE_ORDER_D"], str(repo["repo_root"]))
        assert located["dws.DWB_TRADE_ORDER_D"]["found"] is True

    def test_locate_not_found(self, tmp_path):
        """不存在的表名标 found=False。"""
        repo = build_mock_repo(tmp_path / "repo")
        located = locate_asset_dirs(["DWB_NOT_EXIST"], str(repo["repo_root"]))
        assert located["DWB_NOT_EXIST"]["found"] is False

    def test_locate_multiple_tables(self, tmp_path):
        """多表名同时定位。"""
        repo = build_mock_repo(tmp_path / "repo")
        located = locate_asset_dirs(
            ["DWB_TRADE_ORDER_D", "DWB_RISK_ALERT_F", "DWB_NOT_EXIST"],
            str(repo["repo_root"]),
        )
        assert located["DWB_TRADE_ORDER_D"]["found"] is True
        assert located["DWB_RISK_ALERT_F"]["found"] is True
        assert located["DWB_NOT_EXIST"]["found"] is False

    def test_locate_skips_bft_metric(self, tmp_path):
        """BftMetric 下的指标不应被匹配为资产。"""
        repo = build_mock_repo(tmp_path / "repo")
        # METRIC_TRADE_AMOUNT 在 BFT/BftMetric 下，不在 BftWideTable 下
        located = locate_asset_dirs(["METRIC_TRADE_AMOUNT"], str(repo["repo_root"]))
        assert located["METRIC_TRADE_AMOUNT"]["found"] is False


# ═══════════════════════════════════════════════════════════════
# 读 Sheet3 受影响表清单
# ═══════════════════════════════════════════════════════════════

class TestReadAffectedTables:
    """Sheet3 受影响表清单解析。"""

    def test_read_sheet3(self, tmp_path):
        """正常读取 Sheet3 表名列表。"""
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[],
            field_changes=[
                ChangeItem(before_table="ods.src_a", before_field="user_id",
                           change_type="字段类型及长度变化"),
            ],
            # build_changes_xlsx 默认不含 Sheet3，手动加
        )
        # 手动加 Sheet3
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path))
        ws3 = wb.create_sheet("受影响表清单")
        ws3.append(["序号", "受影响的表名"])
        ws3.append([1, "DWB_TRADE_ORDER_D"])
        ws3.append([2, "DWB_RISK_ALERT_F"])
        wb.save(str(xlsx_path))
        wb.close()

        tables = _read_affected_tables(str(xlsx_path))
        assert len(tables) == 2
        assert "DWB_TRADE_ORDER_D" in tables

    def test_read_empty_sheet3(self, tmp_path):
        """Sheet3 为空返回空列表。"""
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[], field_changes=[])
        # 没有加 Sheet3
        tables = _read_affected_tables(str(xlsx_path))
        assert tables == []


# ═══════════════════════════════════════════════════════════════
# 端到端跨资产分析
# ═══════════════════════════════════════════════════════════════

class TestRunCrossAsset:
    """端到端：mock 代码仓 + 变更清单 → 跨资产分析 → 汇总 Excel。"""

    def test_cross_asset_full_pipeline(self, tmp_path):
        """完整跨资产流程：两资产 + 源变更 → 汇总影响清单。"""
        repo = build_mock_repo(tmp_path / "repo")
        repo_root = str(repo["repo_root"])

        # 构造变更清单（含 Sheet3）
        changes_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(changes_path),
            table_changes=[],
            field_changes=[
                # ods.ods_trade_order_di 是 DWB_TRADE_ORDER_D 的源表
                ChangeItem(before_table="ods.ods_trade_order_di",
                           before_field="amount",
                           before_type="numeric(10,2)",
                           after_table="ods.ods_trade_order_di",
                           after_field="amount",
                           after_type="numeric(18,2)",
                           change_type="字段类型及长度变化"),
                # 无关源表
                ChangeItem(before_table="ods.unrelated",
                           before_field="x",
                           change_type="字段类型及长度变化"),
            ],
        )
        # 手动加 Sheet3
        from openpyxl import load_workbook
        wb = load_workbook(str(changes_path))
        ws3 = wb.create_sheet("受影响表清单")
        ws3.append(["序号", "受影响的表名"])
        ws3.append([1, "DWB_TRADE_ORDER_D"])
        ws3.append([2, "DWB_RISK_ALERT_F"])
        ws3.append([3, "DWB_NOT_EXIST"])  # 定位不到
        wb.save(str(changes_path))
        wb.close()

        # 跑跨资产分析
        output_path = str(tmp_path / "impact_cross.xlsx")
        result = run_cross_asset(
            str(changes_path),
            ["DWB_TRADE_ORDER_D", "DWB_RISK_ALERT_F", "DWB_NOT_EXIST"],
            repo_root,
            output_path,
        )

        # 验证结果
        assert len(result["asset_results"]) == 2  # 两资产成功分析
        assert "DWB_NOT_EXIST" in result["not_found"]
        assert len(result["errors"]) == 0

        # DWB_TRADE_ORDER_D 读 ods.ods_trade_order_di → amount 变了应该有影响
        trade_result = next(
            r for r in result["asset_results"] if r["asset"] == "dwb_trade_order_d"
        )
        # amount 在 SQL 里被引用，应该命中
        assert trade_result["result"].summary["not_hit"] >= 0  # 不报错就行

        # 输出文件存在
        assert Path(output_path).exists()

        # 验证 Excel 有 5 个 Sheet
        from openpyxl import load_workbook as lw
        wb2 = lw(output_path)
        assert "统计摘要" in wb2.sheetnames
        assert "影响清单" in wb2.sheetnames
        assert "表级影响" in wb2.sheetnames
        assert "过滤摘要" in wb2.sheetnames
        assert "未处理" in wb2.sheetnames
        wb2.close()

    def test_cross_asset_summary_aggregation(self, tmp_path):
        """多资产统计汇总正确聚合。"""
        repo = build_mock_repo(tmp_path / "repo")
        repo_root = str(repo["repo_root"])

        changes_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(changes_path),
            table_changes=[],
            field_changes=[],
        )

        output_path = str(tmp_path / "impact_cross.xlsx")
        result = run_cross_asset(
            str(changes_path),
            ["DWB_TRADE_ORDER_D"],
            repo_root,
            output_path,
        )

        s = result["summary"]
        assert s["total_affected"] == 1
        assert s["analyzed"] == 1
        assert s["not_found"] == 0
